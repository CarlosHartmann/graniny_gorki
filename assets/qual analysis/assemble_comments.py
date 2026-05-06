'''
assemble_comments.py: Script to assemble and analyze comments from human and LLM responses.
'''

import csv
import os


def _norm_item_id(item_id):
    item_id = (item_id or '').strip()
    if item_id.startswith('maera'):
        return item_id.replace('maera', 'ma_era_', 1)
    if item_id.startswith('LoRes') and not item_id.startswith('LoRes_'):
        return item_id.replace('LoRes', 'LoRes_', 1)
    return item_id


def _is_source_text(comment, item_id, source_text):
    """Return True if the comment is (a substring of) the original source text for that item."""
    if not comment or item_id not in source_text:
        return False
    return comment.lower() in source_text[item_id]


def _norm_type(value):
    value = (value or '').strip().lower().replace('-', '_').replace(' ', '_')
    value = value.strip("'\",.!?:;()[]{}")
    mapping = {
        'sing': 'singular_they',
        'singular': 'singular_they',
        's': 'singular_they',
        'plural': 'plural_they',
        'pl': 'plural_they',
        'p': 'plural_they',
        'generic': 'generic_they',
        'gen': 'generic_they',
        'g': 'generic_they',
    }
    return mapping.get(value, value)


def assemble_human_comments():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.abspath(os.path.join(script_dir, '..', '..'))

    sheet_paths = [
        os.path.join(
            project_root,
            'raw_data',
            'Humans',
            'linguists_prolific',
            'sheet_version',
            'Block A',
            'block_A_11_5f6ce7a067864711ebb0f66e.csv',
        ),
        os.path.join(
            project_root,
            'raw_data',
            'Humans',
            'linguists_prolific',
            'sheet_version',
            'Block B',
            'Block_B_13_655788501826eeea1b788f8b.csv',
        ),
    ]
    pilotpath = os.path.join(project_root, 'raw_data', 'Humans', 'pilot')
    pilot_sheets = [
        os.path.join(pilotpath, elem)
        for elem in os.listdir(pilotpath)
        if elem.endswith('.csv')
    ]
    sheet_paths.extend(sorted(pilot_sheets))

    limesurvey_path = os.path.join(
        project_root,
        'raw_data',
        'Humans',
        'linguists_prolific',
        'limesurvey',
    )
    limesurvey_files = [
        os.path.join(limesurvey_path, elem)
        for elem in os.listdir(limesurvey_path)
        if elem.endswith('.csv')
    ]

    ground_truth_path = os.path.join(project_root, 'assets', 'ground_truth.csv')
    gt_lookup = {}       # item_id -> canonical they_type
    source_text = {}     # item_id -> original comment_body (lowercased, for contamination check)
    with open(ground_truth_path, newline='', encoding='utf-8-sig') as gt_file:
        for row in csv.DictReader(gt_file):
            raw_id = (row.get('ID') or '').strip()
            if not raw_id:
                continue
            item_id = _norm_item_id(raw_id)
            gt_lookup[item_id] = _norm_type(row.get('they_type'))
            body = (row.get('comment_body') or '').strip().lower()
            if body:
                source_text[item_id] = body

    assembled_comments = {}

    for sheet_file in sheet_paths:
        # the ID is in the column 'ID'
        # annotation is found in column 'they_type' – this can be compared to the correct answers found in "/Users/Carlitos/Documents/GitHub/graniny_gorki/assets/original_annotations/ma_era_wth_lores.xlsx"
        # referent is given in column 'referent' (not always filled in)
        # comment is given in column 'comment' (not always filled in)
        # whether or not an item was context-dependent is given in column 'context_dependency' – 'X' if it was context-dependent, empty otherwise
        # if the respondent had doubts about their answer, they marked the item with 'X' in the column 'doubt'
        # we are going to assemble all received comments per item with each line being formatted like so:
        # [doubt], referent described as [referent], comment: [comment] – [answer in/correct]
        # the first three elements should only be included if they are not empty, and the doubt should be included in square brackets at the beginning of the line if it is present
        # if a respondent had no doubt and did not fill in the referent or comment fields, their response should not be included in the assembled comments for that item
        # the information if they got it right or wrong is only relevant if any of the other three informations are present
        # the assembled comments for each item should be saved in a dictionary with the item ID as the key and the assembled comments as the value
        # this should be returned at the end of the function
        with open(sheet_file, newline='', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            rows = list(reader)[10:]
            for row in rows:
                item_id = _norm_item_id(row.get('ID'))
                if not item_id:
                    continue

                referent = (row.get('referent') or '').strip()
                comment = (row.get('comment') or '').strip()
                doubt = (row.get('doubt') or '').strip()
                answer = _norm_type(row.get('they_type'))
                truth = gt_lookup.get(item_id, '')
                is_correct = answer and truth and answer == truth
                answer_status = 'correct' if is_correct else 'incorrect'

                # Ignore rows with no usable comment metadata.
                if not referent and not comment:
                    continue

                # Skip comments that are just copy-pasted source text.
                if _is_source_text(comment, item_id, source_text):
                    comment = ''
                if not referent and not comment:
                    continue

                line_parts = []
                if doubt:
                    line_parts.append('[doubt]')
                if referent:
                    line_parts.append(f'referent described as {referent}')
                if comment:
                    line_parts.append(f'comment: {comment}')

                assembled_line = ', '.join(line_parts)
                if assembled_line:
                    assembled_line += f' - {answer_status}'
                    assembled_comments.setdefault(item_id, []).append(assembled_line)
    
    for limesurvey_file in limesurvey_files:
        # in these files, each respondent is fully on one row
        # the item IDs are given in column headers, e.g. "maera232684" or "LoRes1690", which would be "ma_era_232684" and "LoRes_1690" in the sheet files
        # the value for each item ID colum is equivalent to "they_type" in the sheet files, and the comment for each item is given in a column with the same name but with "[comment]" appended, e.g. "maera232684[comment]"
        # there are no other metadata columns. So for the assembled lines, we only need this:
        # comment: [comment] – [answer in/correct]
        # skip rows where column "submitdate" is empty, as these are incomplete responses
        with open(limesurvey_file, newline='', encoding='utf-8') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                if not row.get('submitdate'):
                    continue
                for key, value in row.items():
                    if key.endswith('[comment]'):
                        raw_item_id = key[:-9].strip()
                        item_id = _norm_item_id(raw_item_id)  # Remove '[comment]' suffix
                        comment = value.strip()
                        answer = row.get(raw_item_id, '').strip()  # Get the corresponding answer
                        if comment and _is_source_text(comment, item_id, source_text):
                            comment = ''
                        if comment:
                            assembled_line = f'comment: {comment}'
                            answer_norm = _norm_type(answer)
                            truth = gt_lookup.get(item_id, '')
                            is_correct = answer_norm and truth and answer_norm == truth
                            answer_status = 'correct' if is_correct else 'incorrect'
                            assembled_line += f' - {answer_status}'
                            assembled_comments.setdefault(item_id, []).append(assembled_line)

    
    
    return {item_id: '\n'.join(lines) for item_id, lines in assembled_comments.items()}
    

def export_comments(comments_dict):
    # join the assembled comments into a single string for each item with each comment on a new line
    # open the file: /Users/Carlitos/Documents/GitHub/graniny_gorki/assets/annotation_interface/complete_version.xlsm
    # edit its sheet "data"
    # go by column "ID" and overwrite the column "comment" with the assembled comments for each item ID
    # if an item ID from the comments_dict is not found in the sheet, print a warning message
    # save the edited file as a separate xlsx (not xlsm) file in the same directory with the name "assembled_comments.xlsx"
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.abspath(os.path.join(script_dir, '..', '..'))
    annotation_file = os.path.join(
        project_root,
        'assets',
        'annotation_interface',
        'complete_version.xlsm',
    )

    from openpyxl import load_workbook

    wb = load_workbook(annotation_file)
    ws = wb['data']

    id_col_idx = None
    comment_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col_idx).value
        if header_value == 'ID':
            id_col_idx = col_idx
        elif header_value == 'comment':
            comment_col_idx = col_idx

    if id_col_idx is None or comment_col_idx is None:
        print("Error: 'ID' or 'comment' column not found in the sheet.")
        return

    found_ids = set()
    for row_idx in range(2, ws.max_row + 1):
        item_id = ws.cell(row=row_idx, column=id_col_idx).value
        if item_id is None:
            continue
        item_id = str(item_id).strip()
        if item_id in comments_dict:
            ws.cell(row=row_idx, column=comment_col_idx).value = comments_dict[item_id]
            found_ids.add(item_id)

    for missing_id in sorted(set(comments_dict) - found_ids):
        print(f'Warning: item ID not found in workbook: {missing_id}')

    output_file = os.path.join(
        project_root,
        'assets',
        'annotation_interface',
        'assembled_comments_human.xlsx',
    )
    wb.save(output_file)
    print(f'Saved assembled comments to: {output_file}')



if __name__ == '__main__':
    human_comments = assemble_human_comments()
    export_comments(human_comments)